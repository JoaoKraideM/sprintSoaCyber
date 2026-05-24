import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.privacy import criptografar_bytes
from app.core.security import sanitizar_string
from app.services.event_bus import EventBus, EventoDominio
from app.services.veiculo_service import VeiculoService

_FILENAME_REGEX = re.compile(r"[^A-Za-z0-9._-]")
_COLUNA_REGEX = re.compile(r"^coluna\d+$", re.IGNORECASE)
_NUM_INT_REGEX = re.compile(r"^-?\d+$")
_NUM_FLOAT_REGEX = re.compile(r"^-?\d+[.,]\d+$")
_MARCAS_CONHECIDAS = {
    "FORD",
    "TOYOTA",
    "HONDA",
    "CHEVROLET",
    "HYUNDAI",
    "VOLKSWAGEN",
    "NISSAN",
    "JEEP",
    "RENAULT",
    "FIAT",
    "PEUGEOT",
    "CITROEN",
    "BMW",
    "AUDI",
    "MERCEDES-BENZ",
    "MERCEDES",
    "KIA",
}


class ImportacaoExcelError(ValueError):
    def __init__(self, mensagem: str, erros_validacao: list[str] | None = None):
        super().__init__(mensagem)
        self.mensagem = mensagem
        self.erros_validacao = erros_validacao or [mensagem]


class UploadService:
    @staticmethod
    def _upload_dir_seguro() -> Path:
        raiz_projeto = Path(__file__).resolve().parents[2]
        upload_dir = Path(settings.UPLOAD_DIR)
        if not upload_dir.is_absolute():
            upload_dir = (raiz_projeto / upload_dir).resolve()

        raiz_resolvida = raiz_projeto.resolve()
        if raiz_resolvida not in upload_dir.parents and upload_dir != raiz_resolvida:
            raise ValueError("Diretorio de upload fora da raiz do projeto.")

        upload_dir.mkdir(parents=True, exist_ok=True)
        return upload_dir

    @staticmethod
    def _nome_seguro(nome_original: str) -> str:
        nome = Path(nome_original or "").name
        nome = _FILENAME_REGEX.sub("_", nome).strip("._")
        if not nome:
            raise ValueError("Nome de arquivo invalido.")
        return nome

    @staticmethod
    def _validar_tipo_arquivo(nome: str, mime_type: str) -> str:
        extensao = Path(nome).suffix.lower()
        extensoes_permitidas = {ext.lower() for ext in settings.ALLOWED_UPLOAD_EXTENSIONS}
        mimes_permitidos = {mime.lower() for mime in settings.ALLOWED_UPLOAD_CONTENT_TYPES}

        if extensao not in extensoes_permitidas:
            raise ValueError("Somente arquivos Excel (.xlsx ou .xls) sao permitidos.")

        mime = (mime_type or "application/octet-stream").lower()
        if mime not in mimes_permitidos:
            raise ValueError("Tipo MIME nao permitido para upload.")

        return extensao

    @staticmethod
    async def _ler_e_validar_upload(arquivo: UploadFile) -> tuple[str, str, bytes, int]:
        nome_original = UploadService._nome_seguro(arquivo.filename)
        extensao = UploadService._validar_tipo_arquivo(nome_original, arquivo.content_type)

        conteudo = await arquivo.read()
        tamanho = len(conteudo)
        if tamanho == 0:
            raise ValueError("Arquivo vazio nao permitido.")
        if tamanho > settings.MAX_UPLOAD_FILE_SIZE:
            raise ValueError("Arquivo excede o tamanho maximo permitido.")

        return nome_original, extensao, conteudo, tamanho

    @staticmethod
    def _normalizar_texto(valor: Any) -> str:
        if valor is None:
            return ""
        if isinstance(valor, str):
            texto = sanitizar_string(valor)
            return "" if texto.lower() == "nan" else texto
        if isinstance(valor, (int, bool)):
            return str(valor)
        if isinstance(valor, float):
            return str(int(valor)) if valor.is_integer() else str(valor)
        return str(valor).strip()

    @staticmethod
    def _normalizar_valor_metrica(valor: Any) -> Any:
        if valor is None:
            return None
        if isinstance(valor, bool):
            return valor
        if isinstance(valor, int):
            return valor
        if isinstance(valor, float):
            return int(valor) if valor.is_integer() else valor

        texto = UploadService._normalizar_texto(valor)
        if not texto:
            return None

        texto_upper = texto.upper()
        if texto_upper in {"X", "SIM", "TRUE"}:
            return True

        if _NUM_INT_REGEX.fullmatch(texto):
            return int(texto)

        if _NUM_FLOAT_REGEX.fullmatch(texto):
            return float(texto.replace(",", "."))

        return texto

    @staticmethod
    def _valor_ativo(valor: Any) -> bool:
        if valor is None:
            return False
        if isinstance(valor, bool):
            return valor
        if isinstance(valor, (int, float)):
            return valor > 0

        texto = UploadService._normalizar_texto(valor).upper()
        return texto in {"X", "SIM", "TRUE", "1"}

    @staticmethod
    def _valor_int(valor: Any) -> int | None:
        if valor is None:
            return None
        if isinstance(valor, bool):
            return 1 if valor else 0
        if isinstance(valor, int):
            return valor
        if isinstance(valor, float):
            return int(valor) if valor.is_integer() else None

        texto = UploadService._normalizar_texto(valor)
        if not texto:
            return None
        if _NUM_INT_REGEX.fullmatch(texto):
            return int(texto)
        if _NUM_FLOAT_REGEX.fullmatch(texto):
            numero = float(texto.replace(",", "."))
            return int(numero) if numero.is_integer() else None
        return None

    @staticmethod
    def _eh_secao(nome_campo: str, valores_linha: list[Any]) -> bool:
        if not nome_campo:
            return False

        textos = [UploadService._normalizar_texto(valor) for valor in valores_linha]
        preenchidos = [texto for texto in textos if texto]
        if not preenchidos:
            return True

        return all(_COLUNA_REGEX.fullmatch(texto) for texto in preenchidos)

    @staticmethod
    def _detectar_marca_planilha(linhas: list[tuple[Any, ...]]) -> str:
        candidatos = []
        for linha in linhas[:4]:
            for celula in linha:
                texto = UploadService._normalizar_texto(celula).upper()
                if texto:
                    candidatos.append(texto)

        texto_unico = " ".join(candidatos)
        for marca in _MARCAS_CONHECIDAS:
            if re.search(rf"\b{re.escape(marca)}\b", texto_unico):
                return marca
        return ""

    @staticmethod
    def _derivar_motorizacao(atributos: dict[str, Any]) -> str:
        cilindrada_valor = atributos.get("Cilindrada")
        cilindrada_txt = "N/D"
        if isinstance(cilindrada_valor, (int, float)):
            cilindrada_txt = f"{cilindrada_valor:.1f}L"
            if float(cilindrada_valor).is_integer():
                cilindrada_txt = f"{int(cilindrada_valor)}.0L"
        else:
            cilindrada_int = UploadService._valor_int(cilindrada_valor)
            if cilindrada_int is not None:
                cilindrada_txt = f"{cilindrada_int}.0L"

        combustivel = []
        if UploadService._valor_ativo(atributos.get("Motor Diesel")):
            combustivel.append("Diesel")
        if UploadService._valor_ativo(atributos.get("Motor Elétrico")):
            combustivel.append("Eletrico")
        if UploadService._valor_ativo(atributos.get("Motor Flex vs Gasolina")):
            combustivel.append("Flex")

        if not combustivel:
            combustivel.append("Nao informado")

        return f"{cilindrada_txt} {'/'.join(combustivel)}"[:100]

    @staticmethod
    def _derivar_transmissao(atributos: dict[str, Any]) -> str:
        transmissao = "Automatica" if UploadService._valor_ativo(atributos.get("Transmissão Automática")) else "Manual"
        marchas = UploadService._valor_int(atributos.get("Quantidade de marchas"))
        if marchas and marchas > 0:
            transmissao = f"{transmissao} {marchas} marchas"
        return transmissao[:50]

    @staticmethod
    def _derivar_tracao(atributos: dict[str, Any]) -> str:
        if UploadService._valor_ativo(atributos.get("Tração 4x4 (high/low)")):
            return "4x4"
        if UploadService._valor_ativo(atributos.get("Tração integral (AWD)")):
            return "AWD"
        if UploadService._valor_ativo(atributos.get("Tração Traseira")):
            return "Traseira"
        return "Nao informado"

    @staticmethod
    def _parse_planilha_importacao(conteudo: bytes) -> tuple[str, str, list[dict[str, Any]]]:
        erros_validacao: list[str] = []
        try:
            workbook = load_workbook(filename=BytesIO(conteudo), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ImportacaoExcelError(
                "Falha ao ler o arquivo Excel.",
                ["Arquivo invalido ou corrompido para processamento."],
            ) from exc

        if "BASE" not in workbook.sheetnames:
            raise ImportacaoExcelError(
                "Planilha BASE obrigatoria.",
                ["A aba obrigatoria 'BASE' nao foi encontrada no arquivo."],
            )

        worksheet = workbook["BASE"]
        if worksheet.max_row > settings.MAX_EXCEL_ROWS:
            raise ImportacaoExcelError(
                "Planilha excede o limite de linhas.",
                [f"A aba 'BASE' excede o limite de {settings.MAX_EXCEL_ROWS} linhas."],
            )
        if worksheet.max_column > settings.MAX_EXCEL_COLUMNS:
            raise ImportacaoExcelError(
                "Planilha excede o limite de colunas.",
                [f"A aba 'BASE' excede o limite de {settings.MAX_EXCEL_COLUMNS} colunas."],
            )

        linhas = list(worksheet.iter_rows(values_only=True))
        if not linhas:
            raise ImportacaoExcelError(
                "Planilha BASE sem dados.",
                ["A aba 'BASE' esta vazia."],
            )

        cabecalho = linhas[0]
        primeira_coluna = UploadService._normalizar_texto(cabecalho[0] if len(cabecalho) > 0 else "")
        if primeira_coluna != "Equipamentos":
            erros_validacao.append("A primeira coluna deve ser 'Equipamentos'.")

        versoes: list[tuple[int, str]] = []
        for indice, valor in enumerate(cabecalho[1:], start=1):
            nome_versao = UploadService._normalizar_texto(valor)
            if nome_versao:
                versoes.append((indice, nome_versao))

        if not versoes:
            erros_validacao.append("Nenhuma versao encontrada no cabecalho da planilha.")

        modelo = ""
        if len(linhas) > 1:
            linha_modelo = linhas[1]
            for indice, _ in versoes:
                valor = linha_modelo[indice] if indice < len(linha_modelo) else None
                modelo = UploadService._normalizar_texto(valor)
                if modelo:
                    break

        if not modelo:
            erros_validacao.append("Nao foi possivel identificar o modelo na segunda linha da planilha.")

        if erros_validacao:
            raise ImportacaoExcelError("Falha de validacao da planilha.", erros_validacao)

        marca = UploadService._detectar_marca_planilha(linhas) or "FORD"

        metricas_por_versao: dict[str, dict[str, dict[str, Any]]] = {nome: {} for _, nome in versoes}
        atributos_por_versao: dict[str, dict[str, Any]] = {nome: {} for _, nome in versoes}
        secao_atual = "Geral"

        for linha in linhas[2:]:
            if not linha:
                continue
            equipamento = UploadService._normalizar_texto(linha[0] if len(linha) > 0 else "")
            if not equipamento:
                continue

            valores_linha = []
            for indice, _ in versoes:
                valores_linha.append(linha[indice] if indice < len(linha) else None)

            if UploadService._eh_secao(equipamento, valores_linha):
                secao_atual = equipamento
                for _, versao_nome in versoes:
                    metricas_por_versao[versao_nome].setdefault(secao_atual, {})
                continue

            for (indice, versao_nome), valor_bruto in zip(versoes, valores_linha):
                valor = UploadService._normalizar_valor_metrica(valor_bruto)
                metricas_por_versao[versao_nome].setdefault(secao_atual, {})[equipamento] = valor
                atributos_por_versao[versao_nome][equipamento] = valor

        registros_importacao: list[dict[str, Any]] = []
        for _, versao_nome in versoes:
            atributos = atributos_por_versao[versao_nome]
            potencia_cv = UploadService._valor_int(atributos.get("Potência"))
            if potencia_cv is None or potencia_cv <= 0:
                erros_validacao.append(f"Versao '{versao_nome}' sem valor numerico valido para 'Potência'.")
                continue

            registros_importacao.append(
                {
                    "versao": versao_nome,
                    "potencia_cv": potencia_cv,
                    "motorizacao": UploadService._derivar_motorizacao(atributos),
                    "transmissao": UploadService._derivar_transmissao(atributos),
                    "tracao": UploadService._derivar_tracao(atributos),
                    "pacote_equipamentos": metricas_por_versao[versao_nome],
                }
            )

        if erros_validacao:
            raise ImportacaoExcelError("Falha de validacao da planilha.", erros_validacao)

        return marca, modelo, registros_importacao

    @staticmethod
    async def processar_upload_excel(
        db: Session,
        user_id: int,
        user_email: str,
        arquivo: UploadFile,
        ip_origem: str | None = None,
        user_agent: str | None = None,
    ) -> dict:
        nome_original, extensao, conteudo, tamanho = await UploadService._ler_e_validar_upload(arquivo)

        upload_dir = UploadService._upload_dir_seguro()
        sufixo_criptografia = ".enc" if settings.ENCRYPT_UPLOADS_AT_REST else ""
        nome_armazenado = f"{uuid.uuid4().hex}{extensao}{sufixo_criptografia}"
        caminho_arquivo = upload_dir / nome_armazenado
        conteudo_armazenado = criptografar_bytes(conteudo) if settings.ENCRYPT_UPLOADS_AT_REST else conteudo
        caminho_arquivo.write_bytes(conteudo_armazenado)

        EventBus.publicar(
            db,
            EventoDominio(
                nome="ENVIO_INFORMACOES_EXCEL",
                user_id=user_id,
                ip_origem=ip_origem,
                user_agent=user_agent,
                dados_depois={
                    "email": user_email,
                    "tipo_arquivo": "excel",
                    "nome_arquivo": nome_original,
                    "nome_original": nome_original,
                    "nome_armazenado": nome_armazenado,
                    "caminho_armazenado": str(caminho_arquivo),
                    "tamanho_bytes": tamanho,
                    "mime_type": arquivo.content_type or "application/octet-stream",
                    "criptografado_em_repouso": settings.ENCRYPT_UPLOADS_AT_REST,
                    "status_envio": "ARQUIVO_RECEBIDO",
                },
            ),
        )

        return {
            "status": "sucesso",
            "mensagem": "Upload realizado com sucesso.",
            "caminho_arquivo": str(caminho_arquivo),
            "nome_arquivo": nome_original,
            "metrica_id": None,
        }

    @staticmethod
    def publicar_falha_upload(
        db: Session,
        user_id: int,
        erro: str,
        ip_origem: str | None = None,
        user_agent: str | None = None,
    ) -> None:
        EventBus.publicar(
            db,
            EventoDominio(
                nome="FALHA_UPLOAD_EXCEL",
                user_id=user_id,
                ip_origem=ip_origem,
                user_agent=user_agent,
                dados_depois={"erro": erro},
            ),
        )

    @staticmethod
    async def processar_importacao_excel(
        db: Session,
        user_id: int,
        arquivo: UploadFile,
        ip_origem: str | None = None,
        user_agent: str | None = None,
    ) -> dict:
        nome_original, extensao, conteudo, tamanho = await UploadService._ler_e_validar_upload(arquivo)
        if extensao != ".xlsx":
            raise ImportacaoExcelError(
                "Formato nao suportado para importacao estruturada.",
                ["A importacao estruturada aceita apenas arquivos .xlsx."],
            )

        marca, modelo, registros = UploadService._parse_planilha_importacao(conteudo)

        observacao = (
            f"Importacao Excel: {nome_original} em {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC"
        )[:120]

        resultado_catalogo = VeiculoService.importar_registros_excel(
            db=db,
            user_id=user_id,
            marca=marca,
            modelo=modelo,
            registros=registros,
            observacao=observacao,
        )

        for item in resultado_catalogo["itens_importados"]:
            EventBus.publicar(
                db,
                EventoDominio(
                    nome="IMPORTACAO_EXCEL_PROCESSADA",
                    user_id=user_id,
                    metrica_veiculo_id=item["metrica_id"],
                    ip_origem=ip_origem,
                    user_agent=user_agent,
                    dados_depois={
                        "nome_arquivo": nome_original,
                        "marca": marca,
                        "modelo": modelo,
                        "versao": item["versao"],
                        "veiculo_id": item["veiculo_id"],
                        "metrica_id": item["metrica_id"],
                    },
                ),
            )

        if resultado_catalogo["primeira_metrica_id"]:
            EventBus.publicar(
                db,
                EventoDominio(
                    nome="ENVIO_INFORMACOES_EXCEL",
                    user_id=user_id,
                    metrica_veiculo_id=resultado_catalogo["primeira_metrica_id"],
                    ip_origem=ip_origem,
                    user_agent=user_agent,
                    dados_depois={
                        "tipo_arquivo": "datasheet_ford_excel",
                        "nome_arquivo": nome_original,
                        "mime_type": arquivo.content_type or "application/octet-stream",
                        "tamanho_bytes": tamanho,
                        "aba_origem": "BASE",
                        "status_envio": "PROCESSADO",
                        "marca": marca,
                        "modelo": modelo,
                        "versoes_processadas": len(registros),
                        "versoes": resultado_catalogo["versoes_importadas"],
                        "veiculos_criados": resultado_catalogo["veiculos_criados"],
                        "metricas_criadas": resultado_catalogo["metricas_criadas"],
                    },
                ),
            )

        return {
            "status": "sucesso",
            "mensagem": "Processamento de Excel concluido com sucesso.",
            "marca": marca,
            "modelo": modelo,
            "versoes_processadas": len(registros),
            "veiculos_criados": resultado_catalogo["veiculos_criados"],
            "metricas_criadas": resultado_catalogo["metricas_criadas"],
            "erros_validacao": [],
        }
